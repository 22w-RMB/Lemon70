
import requests
import json
import copy
import openpyxl

class readExcel:

    __wb=''
    __Header={
        'X-Lemonban-Media-Type':'lemonban.v2',
        'Content-Type':'application/json'
    }


    def __init__(self,filename):
        self.__wb = openpyxl.load_workbook(filename)

    #包含每一条用例的详细信息的字典
    def __getCaseDict(self,CellList):
        #先用列表存放
        list1 = [
            ['CaseId',''],              #用例编号
            ['CaseModule', ''],          #接口所属模块
            ['CaseTitle', ''],             #用例标题
            ['RequestHeader', ''],          #请求头
            ['RequestWay', ''],         #请求方式
            ['RequestUrl', ''],         #接口地址
            ['RequestBody', ''],        #请求体（参数输入）
            ['Except', ''],             #期望结果
        ]
        for i in range(0,len(list1)):
            list1[i][1] = CellList[i]
        #return的时候转成字典
        return dict(list1)

    #读取每一条用例
    def __readCell(self,SheetName):
        sheet = self.__wb[SheetName]
        #获取列表行列的最大值
        row = sheet.max_row
        col = sheet.max_column
        # cellList用来存放整张表的用例
        Cases = []
        #循环
        for i in range(2,row+1):
            # cellList用来每一条测试用例
            cellList = []
            for j in range(1,col-2):
                #取每个单元格的值
                cell = sheet.cell(row=i,column=j).value
                #因为在excel编写测试用例的参数输入或者预期结果，可能会换行。
                # 而读取的时候会把换行符读取进去，所以需要把换行符\n去掉
                if '\n' in cell:
                    cell = cell.replace('\n',' ')
                #当读取的单元格是参数输入或者预期结果时，需要用eval函数把字符串转成字典
                if j == 7 or j == 8 :
                    cell = eval(cell)
                cellList.append(cell)
            #将每一条测试用例转换成字典格式
            CaseDict = self.__getCaseDict(cellList)
            Cases.append(CaseDict)
        return Cases


    def __requestJudge(self,SheetName,Cases,UserInfo):
        #将请求头深复制
        header = copy.deepcopy(self.__Header)
        #在请求头中增加Authorization
        if UserInfo != None:
            header['Authorization'] = 'Bearer '+UserInfo['token']
        #遍历每一条用例
        for row,CaseDict in enumerate(Cases):
            #获取每条用例的url
            url = CaseDict['RequestUrl']
            #获取每条用例的请求体
            body = CaseDict['RequestBody']
            #如果请求参数存在member_id，则修改
            if 'member_id' in body.keys():
                body['member_id'] = UserInfo['member_id']
            # 获得当前接口的请求方式
            RequestWay = CaseDict['RequestWay'].lower()
            # 接下来判断是那种请求方式
            real_result = ''
            if RequestWay == 'post':
                real_result = self.__PostRequest(url = url,body=body,header=header)
            elif RequestWay == 'get':
                real_result = self.__GetRequest(url = url,header=header)
            elif RequestWay == 'patch':
                real_result = self.__PatchRequest(url = url,body=body,header=header)
            #并将断言结果写入
            self.__write_result(SheetName, CaseDict['Except'], real_result, row+2)

    #Post请求
    def __PostRequest(self,url,body,header):

        real_result = requests.post(url = url,json = body,headers = header).json()
        return real_result

    #Get请求
    def __GetRequest(self,url,header):
        real_result = requests.get(url = url,headers = header).json()
        return real_result

    #Patch请求
    def __PatchRequest(self,url,body,header):
        # print(body)
        real_result = requests.patch(url=url,json = body,headers = header).json()
        return real_result

    #写入数据
    def __write_result(self,SheetName,Except_result,real_result,row):
        sheet = self.__wb[SheetName]
        #实际返回的code
        real_code = real_result['code']
        #预期的code
        except_code =Except_result['code']
        #将返回的结果用json格式输出
        sheet.cell(row=row, column=10).value = json.dumps(real_result,
                                        sort_keys=False,
                                        indent=4,
                                        separators=(' , ',' : '),
                                        ensure_ascii=False)
        #断言
        if real_code == except_code:
            sheet.cell(row=row, column=11).value = "这条用例通过了"
        else:
            sheet.cell(row=row, column=11).value = "这条用例不通过"
        #保存到excel
        self.__wb.save('接口测试用例.xlsx')

    #判断是否需要token，并获取对应的用户信息
    def __JudgeIsNeedToken(self,isNeedToken):
        #不需要token
        if isNeedToken == 0 :
            return None
        #普通用户
        elif isNeedToken == 1 :
            return self.__getUserInfo(1)
        #管理员用户
        elif isNeedToken == 2 :
            return self.__getUserInfo(2)

    #获得token
    def __getUserInfo(self,identity):
        url = 'http://120.78.128.25:8766/futureloan/member/login'
        body = {
        }
        # 普通用户
        if identity == 1:
            body = {
                "mobile_phone": "13548599902",
                "pwd": "78946666"
            }
        # 管理员用户
        elif identity == 2:
            body = {
                "mobile_phone": "13548599901",
                "pwd": "lemon6666"
            }
        result = self.__PostRequest(url,body,self.__Header)
        #将登录后的用户id和token用字典格式return
        return {
            'member_id' : result['data']['id'],
            'token' : result['data']['token_info']['token']
        }

    #开始执行
    def run(self,SheetName,isNeedToken):
        #先获取当前表的所有用例
        Cases = self.__readCell(SheetName)
        #判断是否需要token或者普通用户或者管理员
        UserInfo = self.__JudgeIsNeedToken(isNeedToken)
        #判断请求并执行
        self.__requestJudge(SheetName,Cases,UserInfo)



re =readExcel('接口测试用例.xlsx')
NotNeedToken = 0    #不需要token
user = 1            #普通用户
admin = 2           #管理员

re.run('register',NotNeedToken)
re.run('login',NotNeedToken)
re.run('recharge',user)
re.run('withdraw',user)
re.run('update',user)
re.run('info',admin)
re.run('add',user)
re.run('audit',admin)
re.run('invest',user)
re.run('loans',user)




